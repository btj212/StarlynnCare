#!/usr/bin/env python3
"""
Missouri DHSS Inspections Ingest — missourirecords.xlsx

Parses the FOIA-produced Excel file into inspections + deficiencies rows.

Join key: FACILITY_ID (e.g. "27367N") → strip letter suffix → facilities.external_id
One inspection per (facility_id, EVENT_ID): date=EVENT_DATE, type from SURVEY_CATEGORY.
One deficiency per TAG row (excluding zero-deficiency markers).

Zero-deficiency tags (skip, do not insert deficiency row):
    000L, 00FM, 00IC, 00LC, 000I

Severity mapping (deterministic, no LLM):
  Baseline: 2
  Elevate to 4 when:
    - SURVEY_CATEGORY contains "COMPLAINT INVESTIGATION" (substantiated complaint)
    - TAG description matches abuse/neglect/medication/elopement/evacuation keywords

Usage:
    python3 scrapers/mo_inspections_ingest.py --excel /path/to/missourirecords.xlsx
    python3 scrapers/mo_inspections_ingest.py --excel /path/to/missourirecords.xlsx --dry-run
    python3 scrapers/mo_inspections_ingest.py --excel /path/to/missourirecords.xlsx --state-only
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import psycopg
from psycopg.rows import dict_row

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

EXCEL_DEFAULT = Path(__file__).parent.parent / "data" / "missourirecords.xlsx"

DSN = os.environ.get("SUPABASE_DB_URL") or os.environ.get("DATABASE_URL")

# Zero-deficiency sentinel tags — these rows mean no deficiency was found
ZERO_DEF_TAGS = frozenset({"000L", "00FM", "00IC", "00LC", "000I"})

# Severity-4 keywords in description (case-insensitive)
_SEV4_RE = re.compile(
    r"\b(abuse|neglect|exploit|elopement|medication\s+error|evacuation"
    r"|immediate\s+jeopardy|ij\b|assault|mistreat)\b",
    re.IGNORECASE,
)

# Survey-category → inspection_type mapping
_INSP_TYPE_MAP = {
    "STATE LICENSURE": "routine",
    "INITIAL LICENSURE": "licensure",
    "COMPLAINT INVESTIGATION": "complaint",
    "FOLLOWUP/REVISIT": "revisit",
    "FOLLOWUP/REVISIT, STATE LICENSURE": "revisit",
    "INITIAL LICENSURE, STATE LICENSURE": "licensure",
}

# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------

_INSP_UPSERT = """
INSERT INTO inspections (
    facility_id, inspection_date, inspection_type,
    is_complaint, total_deficiency_count,
    source_url, source_agency, raw_data
)
VALUES (
    %(facility_id)s, %(inspection_date)s, %(inspection_type)s,
    %(is_complaint)s, %(total_deficiency_count)s,
    %(source_url)s, 'MO-DHSS', %(raw_data)s::jsonb
)
ON CONFLICT (facility_id, inspection_date, inspection_type, COALESCE(source_agency, ''))
DO UPDATE SET
    total_deficiency_count = EXCLUDED.total_deficiency_count,
    is_complaint           = EXCLUDED.is_complaint,
    source_url             = EXCLUDED.source_url,
    raw_data               = EXCLUDED.raw_data
RETURNING id
"""

_DEF_INSERT = """
INSERT INTO deficiencies (
    inspection_id, code, description,
    is_repeat, severity, immediate_jeopardy,
    cited_date, state_severity_raw, status
) VALUES (
    %(inspection_id)s, %(code)s, %(description)s,
    %(is_repeat)s, %(severity)s, %(immediate_jeopardy)s,
    %(cited_date)s, %(state_severity_raw)s, 'cited'
)
"""

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _strip_fac_suffix(fid: str | None) -> str | None:
    """'27367N' → '27367' (Socrata facility_number is numeric only)."""
    if not fid:
        return None
    return re.sub(r"[A-Za-z]+$", "", str(fid).strip()) or None


def _parse_date(val: Any) -> date | None:
    if not val:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _map_inspection_type(survey_category: str | None) -> tuple[str, bool]:
    """Return (inspection_type, is_complaint)."""
    raw = (survey_category or "").strip()
    is_complaint = "COMPLAINT INVESTIGATION" in raw.upper()
    # Normalise compound categories to the most specific type
    for key, typ in _INSP_TYPE_MAP.items():
        if key in raw.upper():
            return typ, is_complaint
    return "routine", is_complaint


def _map_severity(survey_category: str | None, description: str | None) -> tuple[int, bool]:
    """Return (severity: int, immediate_jeopardy: bool).

    Severity 4 is reserved for genuine high-harm findings: true immediate
    jeopardy or an abuse/neglect-class keyword match in the description. The
    survey type ("Complaint Investigation") no longer drives severity on its
    own — a routine paperwork finding discovered during a complaint visit is
    not immediate-jeopardy-equivalent. Everything else defaults to severity 2.
    See supabase/migrations/0058_mo_severity_remap.sql.
    """
    desc = (description or "").lower()
    ij = bool(re.search(r"\bimmediate\s+jeopardy\b|\bIJ\b", description or "", re.IGNORECASE))
    if ij:
        return 4, True
    if _SEV4_RE.search(desc):
        return 4, False
    return 2, False


def _source_url(facility_number: str, event_id: Any) -> None:
    # The FOIA Excel has no public deep-link key. The prior scheme
    # (health.mo.gov/seniors/ltcregulation/inspection/<fac>/<event>) was
    # fabricated and 404s, so we emit no per-inspection link here. The real,
    # working deep link (ShowMeLTC inspection_detail.aspx?insid=<insid>) is
    # discovered and written by scrapers/mo_sod_ingest.py during its crawl.
    return None


# ---------------------------------------------------------------------------
# Load Excel
# ---------------------------------------------------------------------------

def load_excel(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed — run: pip3 install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if any(v is not None for v in d.values()):
            rows.append(d)
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Group into inspections → deficiencies
# ---------------------------------------------------------------------------

def group_rows(xl_rows: list[dict[str, Any]]) -> dict[tuple, dict[str, Any]]:
    """
    Group Excel rows by (FACILITY_ID, EVENT_ID) into inspection dicts each
    containing a list of deficiency TAG rows.
    """
    inspections: dict[tuple, dict[str, Any]] = {}
    for r in xl_rows:
        fid = str(r.get("FACILITY_ID") or "").strip()
        eid = str(r.get("EVENT_ID") or "").strip()
        if not fid or not eid:
            continue
        key = (fid, eid)
        if key not in inspections:
            inspections[key] = {
                "facility_id_raw": fid,
                "event_id": eid,
                "event_date": _parse_date(r.get("EVENT_DATE")),
                "survey_category": str(r.get("SURVEY_CATEGORY") or "").strip(),
                "facility_type": str(r.get("FACILITY_TYPE") or "").strip(),
                "deficiency_rows": [],
            }
        tag = str(r.get("TAG") or "").strip().upper()
        if tag not in ZERO_DEF_TAGS and tag:
            inspections[key]["deficiency_rows"].append({
                "tag": tag,
                "description": str(r.get("DESCRIPTION") or "").strip() or None,
            })
    return inspections


# ---------------------------------------------------------------------------
# Main ingest
# ---------------------------------------------------------------------------

def run(excel_path: Path, dry_run: bool) -> None:
    if not DSN:
        sys.exit("Set SUPABASE_DB_URL or DATABASE_URL env var")

    print(f"Loading Excel: {excel_path}", flush=True)
    xl_rows = load_excel(excel_path)
    print(f"  Loaded {len(xl_rows)} rows", flush=True)

    grouped = group_rows(xl_rows)
    print(f"  Grouped into {len(grouped)} inspections", flush=True)

    # Build facility_number → facility_id lookup
    fac_numbers = {_strip_fac_suffix(key[0]) for key in grouped}
    fac_numbers.discard(None)
    print(f"  Unique facility numbers: {len(fac_numbers)}", flush=True)

    fac_map: dict[str, str] = {}  # facility_number → uuid
    with psycopg.connect(DSN, row_factory=dict_row) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT external_id, id FROM facilities WHERE state_code='MO' AND external_id = ANY(%s)",
                (list(fac_numbers),),
            )
            for row in cur.fetchall():
                fac_map[row["external_id"]] = str(row["id"])

    print(f"  Matched {len(fac_map)}/{len(fac_numbers)} facilities in DB", flush=True)

    insp_inserted = insp_skipped = def_inserted = def_errors = 0

    if dry_run:
        print("\nDry-run: counting records that would be written...")
        for (fid_raw, eid), insp in list(grouped.items())[:5]:
            fn = _strip_fac_suffix(fid_raw)
            uid = fac_map.get(fn or "")
            defs = insp["deficiency_rows"]
            print(f"  {fid_raw}/{eid}: date={insp['event_date']} "
                  f"cat={insp['survey_category'][:40]} defs={len(defs)} "
                  f"db_id={'matched' if uid else 'MISSING'}")
        total_defs = sum(len(v["deficiency_rows"]) for v in grouped.values())
        print(f"\nDry-run: {len(grouped)} inspections, {total_defs} deficiencies. No changes written.")
        return

    with psycopg.connect(DSN, row_factory=dict_row, autocommit=True) as conn:
        with conn.cursor() as cur:
            for (fid_raw, eid), insp in grouped.items():
                fn = _strip_fac_suffix(fid_raw)
                facility_id = fac_map.get(fn or "")
                if not facility_id:
                    insp_skipped += 1
                    continue

                insp_type, is_complaint = _map_inspection_type(insp["survey_category"])
                def_rows = insp["deficiency_rows"]
                insp_date = insp["event_date"]

                insp_params = {
                    "facility_id": facility_id,
                    "inspection_date": insp_date,
                    "inspection_type": insp_type,
                    "is_complaint": is_complaint,
                    "total_deficiency_count": len(def_rows),
                    "source_url": _source_url(fn or fid_raw, eid),
                    "raw_data": json.dumps({
                        "event_id": eid,
                        "survey_category": insp["survey_category"],
                        "facility_type": insp["facility_type"],
                    }),
                }

                try:
                    cur.execute(_INSP_UPSERT, insp_params)
                    row = cur.fetchone()
                    if not row:
                        insp_skipped += 1
                        continue
                    insp_id = row["id"]
                    insp_inserted += 1
                except Exception as exc:
                    print(f"  INSP ERROR {fid_raw}/{eid}: {exc}", flush=True)
                    insp_skipped += 1
                    continue

                # Delete-replace deficiencies for this inspection
                cur.execute("DELETE FROM deficiencies WHERE inspection_id = %s", (insp_id,))

                for d in def_rows:
                    sev, ij = _map_severity(insp["survey_category"], d["description"])
                    try:
                        cur.execute(_DEF_INSERT, {
                            "inspection_id": insp_id,
                            "code": d["tag"],
                            "description": d["description"],
                            "is_repeat": False,
                            "severity": sev,
                            "immediate_jeopardy": ij,
                            "cited_date": insp_date,
                            "state_severity_raw": insp["survey_category"],
                        })
                        def_inserted += 1
                    except Exception as exc:
                        print(f"  DEF ERROR {fid_raw}/{eid}/{d['tag']}: {exc}", flush=True)
                        def_errors += 1

                if insp_inserted % 100 == 0:
                    print(f"  {insp_inserted} inspections processed...", end="\r", flush=True)

    print(
        f"\nDone. inspections: inserted={insp_inserted} skipped={insp_skipped}; "
        f"deficiencies: inserted={def_inserted} errors={def_errors}",
        flush=True,
    )
    print("Next: run refresh_snapshot_cache.py --state MO", flush=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Missouri DHSS inspections ingest from Excel")
    parser.add_argument("--excel", default=str(EXCEL_DEFAULT), help="Path to missourirecords.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Print without writing")
    args = parser.parse_args()
    run(excel_path=Path(args.excel), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
