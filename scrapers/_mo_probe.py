#!/usr/bin/env python3
"""
Missouri data probe — confirms Socrata fenu-sipv ↔ Excel join and finalizes
the memory-care signal logic before the full ingest scripts run.

Usage:
    python3 scrapers/_mo_probe.py
    python3 scrapers/_mo_probe.py --excel /path/to/missourirecords.xlsx
"""
import argparse
import json
import os
import sys
from pathlib import Path

import urllib.request
import urllib.parse

SOCRATA_BASE = "https://data.mo.gov/resource/fenu-sipv.json"
EXCEL_DEFAULT = Path(__file__).parent.parent / "data" / "missourirecords.xlsx"

# Memory-care license types (per plan: ALF** = can retain non-self-evacuating residents)
MEMORY_CARE_TYPES = {"ALF**", "ALF"}
# Definitive Alzheimer's SCU flag column name in Socrata
SCU_FLAG_COL = "alzheimer_s_scu"


def fetch_socrata(limit: int = 50, where: str | None = None) -> list[dict]:
    params: dict[str, str | int] = {"$limit": limit}
    if where:
        params["$where"] = where
    url = SOCRATA_BASE + "?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=30) as r:
        return json.loads(r.read())


def run(excel_path: Path) -> None:
    print("=" * 60)
    print("Missouri Data Probe")
    print("=" * 60)

    # ── 1. Socrata sample ──────────────────────────────────────────
    print("\n[1] Fetching 20 rows from Socrata fenu-sipv...")
    rows = fetch_socrata(limit=20)
    if not rows:
        print("ERROR: No data returned from Socrata")
        sys.exit(1)

    print(f"    Columns: {list(rows[0].keys())}")
    print(f"    Sample facility_number: {rows[0].get('facility_number')}")
    print(f"    Sample level_of_care:   {rows[0].get('level_of_care')}")
    print(f"    Sample alzheimer_s_scu: {rows[0].get(SCU_FLAG_COL)}")
    print(f"    Sample name:            {rows[0].get('entity_name') or rows[0].get('name')}")

    # ── 2. Memory-care set size ───────────────────────────────────
    print("\n[2] Counting ALF** / alzheimer_s_scu rows in Socrata...")
    all_rows = fetch_socrata(limit=5000)
    alf_star = [r for r in all_rows if (r.get("level_of_care") or "").strip().upper() in {"ALF**", "ALF"}]
    scu_true = [r for r in all_rows if str(r.get(SCU_FLAG_COL, "")).lower() == "true"]
    combined = {r.get("facility_number") for r in alf_star} | {r.get("facility_number") for r in scu_true}
    print(f"    Total rows:             {len(all_rows)}")
    print(f"    ALF/ALF** rows:         {len(alf_star)}")
    print(f"    alzheimer_s_scu=true:   {len(scu_true)}")
    print(f"    Union (unique fac#):    {len(combined)}")

    # License status check
    license_cols = [k for k in (rows[0].keys() if rows else []) if "licens" in k.lower() or "expir" in k.lower()]
    print(f"\n[3] License/expiration columns: {license_cols}")
    sample_alf = next((r for r in alf_star[:5]), None)
    if sample_alf:
        for col in license_cols:
            print(f"    {col}: {sample_alf.get(col)}")

    # ── 4. Excel join ─────────────────────────────────────────────
    if not excel_path.exists():
        print(f"\n[4] Excel not found at {excel_path} — skipping join test")
        return

    try:
        import openpyxl
    except ImportError:
        print("\n[4] openpyxl not installed — run: pip3 install openpyxl")
        print("    Skipping Excel join test")
        return

    print(f"\n[4] Loading Excel from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"    Excel columns: {headers}")

    rows_xl = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        rows_xl.append(d)
    wb.close()

    print(f"    Excel rows: {len(rows_xl)}")
    if rows_xl:
        print(f"    Sample FACILITY_ID: {rows_xl[0].get('FACILITY_ID')}")
        print(f"    Sample FACILITY_TYPE: {rows_xl[0].get('FACILITY_TYPE')}")
        print(f"    Sample TAG: {rows_xl[0].get('TAG')}")

    # Derive numeric facility number from Excel FACILITY_ID (e.g. "27367N" → "27367")
    def strip_suffix(fid: str | None) -> str | None:
        if not fid:
            return None
        return str(fid).rstrip("NABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz")

    xl_fac_numbers = {strip_suffix(r.get("FACILITY_ID")) for r in rows_xl if r.get("FACILITY_ID")}
    socrata_fac_numbers = {str(r.get("facility_number", "")).strip() for r in all_rows if r.get("facility_number")}

    matched = xl_fac_numbers & socrata_fac_numbers
    unmatched = xl_fac_numbers - socrata_fac_numbers
    print(f"\n[5] Join check (facility_number):")
    print(f"    Excel unique IDs:       {len(xl_fac_numbers)}")
    print(f"    Socrata unique IDs:     {len(socrata_fac_numbers)}")
    print(f"    Matched:                {len(matched)}")
    print(f"    Unmatched Excel IDs:    {len(unmatched)}")
    if unmatched and len(unmatched) <= 10:
        print(f"    Unmatched sample:       {sorted(unmatched)[:10]}")

    # ── 6. Distinct tags in Excel (for severity mapping) ─────────
    tags = {str(r.get("TAG", "")).strip() for r in rows_xl if r.get("TAG")}
    zero_def_tags = {"000L", "00FM", "00IC", "00LC", "000I"}
    actual_tags = tags - zero_def_tags
    print(f"\n[6] Distinct TAGs in Excel:   {len(tags)}")
    print(f"    Zero-deficiency tags:     {zero_def_tags & tags}")
    print(f"    Substantive tags:         {len(actual_tags)}")
    # Show survey categories
    cats = {str(r.get("SURVEY_CATEGORY", "")) for r in rows_xl if r.get("SURVEY_CATEGORY")}
    print(f"    SURVEY_CATEGORY values:   {sorted(cats)}")

    print("\n" + "=" * 60)
    print("Probe complete.")
    print("=" * 60)


def main() -> None:
    parser = argparse.ArgumentParser(description="Missouri data probe")
    parser.add_argument("--excel", default=str(EXCEL_DEFAULT), help="Path to missourirecords.xlsx")
    args = parser.parse_args()
    run(excel_path=Path(args.excel))


if __name__ == "__main__":
    main()
