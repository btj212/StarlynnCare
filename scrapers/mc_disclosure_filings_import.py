#!/usr/bin/env python3
"""
Strategy B — Import CDSS §1569.627 memory-care disclosure filings (PRA / bulk export).

Reads a tabular file from CDSS (CSV or XLSX). Format details are finalized when the
data arrives; this script uses flexible column detection (see LICENSE_COLUMNS, etc.).

Writes (idempotent, never sets disclosure to false):
  memory_care_disclosure_filed = true
  memory_care_disclosure_source = 'cdss_pra_2026'

Match order:
  1. Exact match on normalized license_number (9-digit zero-padded string).
  2. Fuzzy fallback on (name, city, zip) against CA licensed facilities — same city/zip
     filter + RapidFuzz name score (see --fuzzy-threshold).

Post-import
-----------
Run after this script completes successfully:

    python3 scrapers/recompute_publishable.py

That promotes eligible ``needs_review`` rows once the Tier-1 disclosure signal exists.

Usage
-----
    python3 scrapers/mc_disclosure_filings_import.py --input path/to/file.csv --dry-run
    python3 scrapers/mc_disclosure_filings_import.py --input disclosures.xlsx --format xlsx
    python3 scrapers/mc_disclosure_filings_import.py --input data.csv --unmatched-out /tmp/missing.csv
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path
from typing import Any

import psycopg
from dotenv import load_dotenv

try:
    from rapidfuzz import fuzz
except ImportError:
    print("rapidfuzz not installed. pip install rapidfuzz>=3.0.0", file=sys.stderr)
    sys.exit(2)

REPO_ROOT = Path(__file__).resolve().parent.parent
STATE_CODE = "CA"
SOURCE_TAG = "cdss_pra_2026"

# Header synonyms → canonical keys (lowercase comparison).
LICENSE_COLUMNS = (
    "license_number",
    "license",
    "facility_number",
    "facility number",
    "facility_no",
    "lic_no",
    "lic #",
    "license no",
    "license #",
)
NAME_COLUMNS = ("name", "facility_name", "facility name", "business_name", "dba")
CITY_COLUMNS = ("city", "facility_city", "mail_city")
ZIP_COLUMNS = ("zip", "zip_code", "zip code", "postal", "postal_code")


def load_env() -> None:
    for name in (".env.local", ".env"):
        p = REPO_ROOT / name
        if p.is_file():
            load_dotenv(p)


def normalize_license(raw: str | None) -> str | None:
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw).strip())
    if not digits:
        return None
    if len(digits) > 9:
        digits = digits[-9:]
    return digits.zfill(9)


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _pick_columns(fieldnames: list[str]) -> dict[str, str | None]:
    """Map canonical keys license, name, city, zip to actual header strings."""
    norm_map = {_norm_header(f): f for f in fieldnames}
    out: dict[str, str | None] = {
        "license": None,
        "name": None,
        "city": None,
        "zip": None,
    }
    for cand in LICENSE_COLUMNS:
        k = _norm_header(cand)
        if k in norm_map:
            out["license"] = norm_map[k]
            break
    for cand in NAME_COLUMNS:
        k = _norm_header(cand)
        if k in norm_map:
            out["name"] = norm_map[k]
            break
    for cand in CITY_COLUMNS:
        k = _norm_header(cand)
        if k in norm_map:
            out["city"] = norm_map[k]
            break
    for cand in ZIP_COLUMNS:
        k = _norm_header(cand)
        if k in norm_map:
            out["zip"] = norm_map[k]
            break
    return out


def _normalize_zip(z: str | None) -> str:
    if not z:
        return ""
    d = re.sub(r"\D", "", str(z))
    if len(d) >= 5:
        return d[:5]
    return d


def _normalize_name(name: str | None) -> str:
    if not name:
        return ""
    s = name.lower()
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def load_rows_csv(path: Path) -> tuple[list[dict[str, str]], dict[str, str | None]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row")
        cols = _pick_columns(list(reader.fieldnames))
        rows: list[dict[str, str]] = []
        for r in reader:
            rows.append({k: (v or "").strip() if v is not None else "" for k, v in r.items()})
    return rows, cols


def load_rows_xlsx(path: Path) -> tuple[list[dict[str, str]], dict[str, str | None]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        print(
            "openpyxl required for .xlsx. pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(2)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("XLSX is empty")

    fieldnames = [str(c or "").strip() for c in header_row]
    cols = _pick_columns(fieldnames)
    body: list[dict[str, str]] = []
    for row in rows_iter:
        d: dict[str, str] = {}
        for i, fn in enumerate(fieldnames):
            val = row[i] if i < len(row) else None
            if val is None:
                d[fn] = ""
            else:
                d[fn] = str(val).strip()
        if any(v.strip() for v in d.values()):
            body.append(d)
    wb.close()
    return body, cols


def extract_records(
    rows: list[dict[str, str]],
    cols: dict[str, str | None],
) -> list[dict[str, str | None]]:
    """Each record: license_norm, name, city, zip_raw."""
    lic_h = cols.get("license")
    name_h = cols.get("name")
    city_h = cols.get("city")
    zip_h = cols.get("zip")
    if not lic_h and not (name_h and city_h):
        raise ValueError(
            "Could not detect columns. Need a license column and/or name+city. "
            f"Headers seen: {sorted({k for r in rows[:1] for k in r.keys()})}"
        )
    out: list[dict[str, str | None]] = []
    for r in rows:
        lic = normalize_license(r.get(lic_h, "") if lic_h else "") if lic_h else None
        name = (r.get(name_h, "") if name_h else "").strip() or None
        city = (r.get(city_h, "") if city_h else "").strip() or None
        zraw = (r.get(zip_h, "") if zip_h else "").strip() or None
        out.append(
            {
                "license_norm": lic,
                "name": name,
                "city": city,
                "zip": zraw,
            }
        )
    return out


def fetch_facility_maps(
    conn: psycopg.Connection,
) -> tuple[dict[str, str], list[tuple[str, str, str, str, str]]]:
    """
    Returns:
      by_license: license -> facility_id
      all_rows: list of (id, name, city, zip, license_number) for fuzzy matching
    """
    sql = """
        SELECT id::text, name, COALESCE(city, ''), COALESCE(zip, ''),
               COALESCE(license_number, '')
        FROM facilities
        WHERE state_code = %s AND license_status = 'LICENSED'
    """
    by_license: dict[str, str] = {}
    all_rows: list[tuple[str, str, str, str, str]] = []
    with conn.cursor() as cur:
        cur.execute(sql, (STATE_CODE,))
        for fid, name, city, zipc, lic in cur.fetchall():
            ln = normalize_license(lic)
            if ln:
                by_license[ln] = fid
            all_rows.append((fid, name, city, zipc, lic))
    return by_license, all_rows


def fuzzy_match_facility(
    name: str | None,
    city: str | None,
    zip_raw: str | None,
    all_rows: list[tuple[str, str, str, str, str]],
    threshold: int,
) -> tuple[str, int] | None:
    """Return (facility_id, score) or None."""
    if not name or not city:
        return None
    z = _normalize_zip(zip_raw)
    target_name = _normalize_name(name)
    if not target_name:
        return None
    city_l = city.strip().lower()
    best_id: str | None = None
    best_score = 0
    for fid, fname, fcity, fzip, _ in all_rows:
        if fcity.strip().lower() != city_l:
            continue
        if z and _normalize_zip(fzip) and _normalize_zip(fzip) != z:
            continue
        score = fuzz.token_set_ratio(target_name, _normalize_name(fname))
        if score > best_score:
            best_score = score
            best_id = fid
    if best_id and best_score >= threshold:
        return (best_id, best_score)
    return None


def apply_updates(
    conn: psycopg.Connection,
    facility_ids: list[str],
    dry_run: bool,
) -> int:
    """Set disclosure flags; returns rows updated (or would update)."""
    if not facility_ids:
        return 0
    sql = """
        UPDATE facilities
        SET
          memory_care_disclosure_filed = true,
          memory_care_disclosure_source = %s
        WHERE id = %s::uuid
          AND state_code = %s
          AND (
            memory_care_disclosure_filed = false
            OR memory_care_disclosure_source IS DISTINCT FROM %s
          )
    """
    updated = 0
    with conn.cursor() as cur:
        for fid in facility_ids:
            if dry_run:
                cur.execute(
                    "SELECT memory_care_disclosure_filed, memory_care_disclosure_source "
                    "FROM facilities WHERE id = %s::uuid",
                    (fid,),
                )
                row = cur.fetchone()
                if row:
                    would_change = (not row[0]) or (row[1] != SOURCE_TAG)
                    if would_change:
                        print(
                            f"  [dry-run] facility_id={fid} "
                            f"current filed={row[0]} source={row[1]!r} → would set "
                            f"filed=true source={SOURCE_TAG!r}"
                        )
                        updated += 1
                continue
            cur.execute(sql, (SOURCE_TAG, fid, STATE_CODE, SOURCE_TAG))
            updated += cur.rowcount
    return updated


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import §1569.627 disclosure filings list → memory_care_disclosure_filed",
    )
    parser.add_argument("--input", required=True, help="Path to CSV or XLSX")
    parser.add_argument(
        "--format",
        choices=("auto", "csv", "xlsx"),
        default="auto",
        help="File format (default: infer from extension)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print matches; no DB writes")
    parser.add_argument(
        "--unmatched-out",
        help="Write unmatched input rows to this CSV path",
    )
    parser.add_argument(
        "--fuzzy-threshold",
        type=int,
        default=85,
        help="RapidFuzz token_set_ratio minimum for name fallback (default: 85)",
    )
    parser.add_argument(
        "--max-unmatched-print",
        type=int,
        default=25,
        help="Max unmatched rows to print as sample (default: 25)",
    )
    args = parser.parse_args()

    path = Path(args.input)
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    fmt = args.format
    if fmt == "auto":
        suf = path.suffix.lower()
        if suf in (".csv", ".tsv"):
            fmt = "csv"
        elif suf in (".xlsx", ".xlsm"):
            fmt = "xlsx"
        else:
            print(
                "Could not infer format; use --format csv or --format xlsx",
                file=sys.stderr,
            )
            return 1

    load_env()
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("DATABASE_URL not set (required for facility matching)", file=sys.stderr)
        return 1

    if fmt == "csv":
        rows, cols = load_rows_csv(path)
    else:
        rows, cols = load_rows_xlsx(path)

    print(f"Loaded {len(rows)} data rows from {path.name}")
    print(f"Column mapping: {cols}")

    try:
        records = extract_records(rows, cols)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    # Deduplicate by license when present; else by (name, city, zip)
    seen: set[str] = set()
    deduped: list[dict[str, str | None]] = []
    for rec in records:
        lic_part = rec["license_norm"] or ""
        key = lic_part or f"{rec['name']}|{rec['city']}|{rec['zip']}"
        if key in seen:
            continue
        seen.add(key)
        deduped.append(rec)

    matched_license: list[tuple[str, str]] = []  # (facility_id, license)
    matched_fuzzy: list[tuple[str, str, int]] = []  # (facility_id, label, score)
    unmatched: list[dict[str, str | None]] = []

    with psycopg.connect(dsn) as conn:
        by_license, all_rows = fetch_facility_maps(conn)

        for rec in deduped:
            lic = rec["license_norm"]
            fid: str | None = None
            if lic and lic in by_license:
                fid = by_license[lic]
            if fid:
                matched_license.append((fid, lic or ""))
                continue

            fm = fuzzy_match_facility(
                rec["name"],
                rec["city"],
                rec["zip"],
                all_rows,
                args.fuzzy_threshold,
            )
            if fm:
                fid, score = fm
                matched_fuzzy.append((fid, rec["name"] or "", score))
                continue

            unmatched.append(rec)

        id_by_license = {fid for fid, _ in matched_license}
        combined_ids: list[str] = []
        for fid, _ in matched_license:
            combined_ids.append(fid)
        for fid, _, _ in matched_fuzzy:
            if fid not in id_by_license:
                combined_ids.append(fid)
        # Preserve order, unique
        seen_ids: set[str] = set()
        unique_ids: list[str] = []
        for fid in combined_ids:
            if fid not in seen_ids:
                seen_ids.add(fid)
                unique_ids.append(fid)

        print(f"\nMatch summary:")
        print(f"  By license: {len(matched_license)}")
        print(f"  By fuzzy name+city+zip: {len(matched_fuzzy)}")
        print(f"  Unmatched: {len(unmatched)}")
        print(f"  Unique facility rows to update: {len(unique_ids)}")

        if unmatched and args.max_unmatched_print > 0:
            print(f"\nUnmatched sample (up to {args.max_unmatched_print}):")
            for rec in unmatched[: args.max_unmatched_print]:
                print(
                    f"  lic={rec['license_norm']!r} name={rec['name']!r} "
                    f"city={rec['city']!r} zip={rec['zip']!r}"
                )

        if args.unmatched_out and unmatched:
            outp = Path(args.unmatched_out)
            outp.parent.mkdir(parents=True, exist_ok=True)
            with outp.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(
                    f,
                    fieldnames=["license_norm", "name", "city", "zip"],
                )
                w.writeheader()
                for rec in unmatched:
                    w.writerow(
                        {
                            "license_norm": rec["license_norm"] or "",
                            "name": rec["name"] or "",
                            "city": rec["city"] or "",
                            "zip": rec["zip"] or "",
                        }
                    )
            print(f"Wrote {len(unmatched)} unmatched rows to {outp}")

        n = apply_updates(conn, unique_ids, args.dry_run)
        if not args.dry_run:
            conn.commit()
        print(f"\nRows updated (or distinct source changed): {n}")

    if args.dry_run:
        print("\nDry-run complete. No database changes committed.")
    else:
        print("\nDone. Run: python3 scrapers/recompute_publishable.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
