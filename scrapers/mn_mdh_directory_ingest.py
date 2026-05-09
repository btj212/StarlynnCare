#!/usr/bin/env python3
"""
Ingest Minnesota MDH facility directory Excel → facilities.

Filters Assisted Living providers (`ALL_PROV = Y`) and flags
`mn_dementia_care_licensed` when LIC_TYPE / HCP_TYPE indicates 144G dementia care
(substring match on field text — confirm codes against MDH documentation in
[`docs/MN_DATA_SOURCES.md`](../docs/MN_DATA_SOURCES.md)).

Optional `--name-fallback`: also flag facilities whose name contains dementia/memory
keywords (weaker signal — off by default).

Usage:
  python3 scrapers/mn_mdh_directory_ingest.py --input .firecrawl/mn-scrape/facility-directory.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path
from typing import Any

import psycopg
from dotenv import load_dotenv

REPO_ROOT = Path(__file__).resolve().parent.parent
STATE_CODE = "MN"
SOURCE_PAGE = "https://www.health.state.mn.us/facilities/regulation/directory/directorydatafile.html"


def load_env() -> None:
    for name in (".env.local", ".env"):
        p = REPO_ROOT / name
        if p.is_file():
            load_dotenv(p)


def slugify(text: str) -> str:
    s = text.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "facility"


def titleize(text: str) -> str:
    if not text:
        return text
    stop = {"and", "or", "of", "in", "the", "a", "at", "by", "for", "to"}
    words = text.lower().split()
    result = []
    for i, w in enumerate(words):
        result.append(w if (i > 0 and w in stop) else w.capitalize())
    return " ".join(result)


def pad_mn_hfid(raw: Any) -> str:
    if raw is None or str(raw).strip() == "":
        return "0000000000"
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return "0000000000"
    return digits.zfill(10)


def facility_slug(name: str, lic: str) -> str:
    base = slugify(name)
    suffix = lic[-8:].lstrip("0") or lic[-4:]
    return f"{base}-{suffix}"


def norm_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def cell(row: tuple[Any, ...], idx: dict[str, int], *names: str) -> Any:
    for n in names:
        k = norm_header(n)
        if k in idx and idx[k] < len(row):
            return row[idx[k]]
    return None


def infer_dementia_lic(
    lic_type: str | None,
    hcp_type: str | None,
    *,
    name_fallback: bool,
    name: str | None,
) -> bool:
    lt = (lic_type or "").lower()
    ht = (hcp_type or "").lower()
    if "dementia" in lt or "dementia" in ht:
        return True
    if "144g" in lt or "144g" in ht:
        return True
    if "memory care" in lt or "memory care" in ht:
        return True
    if name_fallback and name:
        nm = name.lower()
        for kw in ("dementia", "memory care", "alzheimer"):
            if kw in nm:
                return True
    return False


def find_header_row(ws: Any) -> tuple[int, list[str]]:
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=30)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(norm_header(v) == "hfid" or "hfid" in norm_header(v) for v in vals if v):
            return i, vals
        if any("health facility id" in norm_header(v) for v in vals if v):
            return i, vals
    raise ValueError("Could not find header row with HFID")


def build_header_map(header_cells: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for j, cell in enumerate(header_cells):
        key = norm_header(cell)
        if key:
            idx[key] = j
    return idx


MN_UPSERT_SQL = """
INSERT INTO facilities (
    state_code, name, cms_id,
    license_number, license_type,
    street, city, zip,
    city_slug, slug,
    beds, facility_type, certification_type,
    operator_name, management_company, ownership_type,
    phone, website,
    cms_star_rating, last_inspection_date,
    latitude, longitude,
    source_url,
    care_category, serves_memory_care, memory_care_designation,
    license_status, license_expiration, publishable,
    mc_signal_explicit_name, mc_signal_chain_name, mc_review_status,
    memory_care_disclosure_filed, memory_care_disclosure_source,
    mc_signal_apfm_listed, mc_signal_caring_listed,
    tx_license_class, tx_alzheimer_certified,
    tx_facility_id, tx_alzheimer_capacity, tx_alzheimer_cert_no,
    tx_alzheimer_cert_effective, tx_alzheimer_cert_expiration,
    tx_license_effective, tx_license_expiration, tx_license_initial,
    tx_state_region, tx_hhsc_suboffice, tx_county,
    mn_dementia_care_licensed, mn_hfid, mn_lic_type, mn_hcp_type,
    mn_all_capacity, mn_county_code,
    updated_at
) VALUES (
    %(state_code)s, %(name)s, %(cms_id)s,
    %(license_number)s, %(license_type)s,
    %(street)s, %(city)s, %(zip)s,
    %(city_slug)s, %(slug)s,
    %(beds)s, %(facility_type)s, %(certification_type)s,
    %(operator_name)s, %(management_company)s, %(ownership_type)s,
    %(phone)s, %(website)s,
    %(cms_star_rating)s, %(last_inspection_date)s,
    %(latitude)s, %(longitude)s,
    %(source_url)s,
    %(care_category)s, %(serves_memory_care)s, %(memory_care_designation)s,
    %(license_status)s, %(license_expiration)s, %(publishable)s,
    %(mc_signal_explicit_name)s, %(mc_signal_chain_name)s, %(mc_review_status)s,
    %(memory_care_disclosure_filed)s, %(memory_care_disclosure_source)s,
    %(mc_signal_apfm_listed)s, %(mc_signal_caring_listed)s,
    %(tx_license_class)s, %(tx_alzheimer_certified)s,
    %(tx_facility_id)s, %(tx_alzheimer_capacity)s, %(tx_alzheimer_cert_no)s,
    %(tx_alzheimer_cert_effective)s, %(tx_alzheimer_cert_expiration)s,
    %(tx_license_effective)s, %(tx_license_expiration)s, %(tx_license_initial)s,
    %(tx_state_region)s, %(tx_hhsc_suboffice)s, %(tx_county)s,
    %(mn_dementia_care_licensed)s, %(mn_hfid)s, %(mn_lic_type)s, %(mn_hcp_type)s,
    %(mn_all_capacity)s, %(mn_county_code)s,
    now()
)
ON CONFLICT (state_code, city_slug, slug) DO UPDATE SET
    name                         = EXCLUDED.name,
    license_number               = EXCLUDED.license_number,
    license_type                 = EXCLUDED.license_type,
    street                       = EXCLUDED.street,
    city                         = EXCLUDED.city,
    zip                          = EXCLUDED.zip,
    beds                         = EXCLUDED.beds,
    facility_type                = EXCLUDED.facility_type,
    source_url                   = EXCLUDED.source_url,
    care_category                = EXCLUDED.care_category,
    serves_memory_care           = EXCLUDED.serves_memory_care,
    memory_care_designation      = EXCLUDED.memory_care_designation,
    license_status               = EXCLUDED.license_status,
    publishable                  = EXCLUDED.publishable,
    mc_review_status             = CASE
        WHEN facilities.mc_review_status IN ('reviewed_publish','reviewed_reject')
        THEN facilities.mc_review_status
        ELSE EXCLUDED.mc_review_status
    END,
    mn_dementia_care_licensed    = EXCLUDED.mn_dementia_care_licensed,
    mn_hfid                      = EXCLUDED.mn_hfid,
    mn_lic_type                  = EXCLUDED.mn_lic_type,
    mn_hcp_type                  = EXCLUDED.mn_hcp_type,
    mn_all_capacity              = EXCLUDED.mn_all_capacity,
    mn_county_code               = EXCLUDED.mn_county_code,
    updated_at                   = now()
"""


def build_row(
    raw_row: tuple[Any, ...],
    idx: dict[str, int],
    *,
    name_fallback: bool,
) -> dict[str, Any] | None:
    all_prov = cell(raw_row, idx, "all_prov", "all prov")
    if all_prov is None or str(all_prov).strip().upper() != "Y":
        return None

    hfid_raw = cell(raw_row, idx, "hfid", "health facility id")
    name_raw = cell(raw_row, idx, "name")
    if hfid_raw is None or name_raw is None:
        return None
    hfid = str(hfid_raw).strip()
    raw_name = str(name_raw).strip()
    if not hfid or not raw_name:
        return None

    lic = pad_mn_hfid(hfid)
    name = titleize(raw_name) or f"Facility {lic}"

    lic_type = str(cell(raw_row, idx, "lic_type") or "").strip() or None
    hcp_type = str(cell(raw_row, idx, "hcp_type") or "").strip() or None

    dc = infer_dementia_lic(
        lic_type,
        hcp_type,
        name_fallback=name_fallback,
        name=raw_name,
    )

    raw_city = str(cell(raw_row, idx, "city") or "").strip()
    city_slug = slugify(raw_city) if raw_city else "unknown-city"
    city = titleize(raw_city) if raw_city else None

    street = titleize(str(cell(raw_row, idx, "address") or "").strip()) or None
    zip_raw = cell(raw_row, idx, "zip")
    zip_str = str(zip_raw).strip()[:10] if zip_raw else None

    cap_raw = cell(raw_row, idx, "all_capacity", "all capacity")
    cap: int | None = None
    if cap_raw is not None and str(cap_raw).strip() != "":
        try:
            cap = int(float(str(cap_raw)))
        except (ValueError, TypeError):
            cap = None

    county_code = str(cell(raw_row, idx, "county_code", "county code") or "").strip() or None

    slug = facility_slug(name, lic)

    return {
        "state_code": STATE_CODE,
        "name": name,
        "cms_id": None,
        "license_number": lic,
        "license_type": lic_type,
        "street": street,
        "city": city,
        "zip": zip_str,
        "city_slug": city_slug,
        "slug": slug,
        "beds": cap,
        "facility_type": "alf",
        "certification_type": "state",
        "operator_name": None,
        "management_company": None,
        "ownership_type": None,
        "phone": str(cell(raw_row, idx, "telephone", "phone") or "").strip() or None,
        "website": None,
        "cms_star_rating": None,
        "last_inspection_date": None,
        "latitude": None,
        "longitude": None,
        "source_url": SOURCE_PAGE,
        "care_category": "alf_memory_care" if dc else "alf_general",
        "serves_memory_care": dc,
        "memory_care_designation": (
            "Minnesota Assisted Living with Dementia Care (144G)" if dc else None
        ),
        "license_status": "LICENSED",
        "license_expiration": None,
        "publishable": False,
        "mc_signal_explicit_name": False,
        "mc_signal_chain_name": False,
        "mc_review_status": "auto_published",
        "memory_care_disclosure_filed": False,
        "memory_care_disclosure_source": None,
        "mc_signal_apfm_listed": False,
        "mc_signal_caring_listed": False,
        "tx_license_class": None,
        "tx_alzheimer_certified": False,
        "tx_facility_id": None,
        "tx_alzheimer_capacity": None,
        "tx_alzheimer_cert_no": None,
        "tx_alzheimer_cert_effective": None,
        "tx_alzheimer_cert_expiration": None,
        "tx_license_effective": None,
        "tx_license_expiration": None,
        "tx_license_initial": None,
        "tx_state_region": None,
        "tx_hhsc_suboffice": None,
        "tx_county": None,
        "mn_dementia_care_licensed": dc,
        "mn_hfid": hfid,
        "mn_lic_type": lic_type,
        "mn_hcp_type": hcp_type,
        "mn_all_capacity": cap,
        "mn_county_code": county_code,
    }


def upsert_facilities(conn: psycopg.Connection, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with conn.cursor() as cur:
        cur.executemany(MN_UPSERT_SQL, rows)


def main() -> int:
    ap = argparse.ArgumentParser(description="MDH directory Excel → facilities")
    ap.add_argument("--input", type=Path, required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--name-fallback",
        action="store_true",
        help="Flag dementia from facility name keywords (weaker).",
    )
    args = ap.parse_args()
    load_env()

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        print("Install openpyxl", file=sys.stderr)
        return 2

    if not args.input.is_file():
        print(f"Not found: {args.input}", file=sys.stderr)
        return 1

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header_row_idx, header_cells = find_header_row(ws)
    header_map = build_header_map(header_cells)

    built: list[dict[str, Any]] = []
    dc_count = 0
    for raw_row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        if not raw_row or not any(raw_row):
            continue
        rec = build_row(raw_row, header_map, name_fallback=args.name_fallback)
        if rec is None:
            continue
        if rec["mn_dementia_care_licensed"]:
            dc_count += 1
        built.append(rec)
    wb.close()

    print(f"Rows upsert-ready: {len(built)} (mn_dementia_care_licensed={dc_count})")

    if args.dry_run:
        return 0

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("DATABASE_URL not set", file=sys.stderr)
        return 1

    with psycopg.connect(dsn) as conn:
        upsert_facilities(conn, built)
        conn.commit()
    print(f"Upserted {len(built)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
