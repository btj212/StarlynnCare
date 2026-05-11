#!/usr/bin/env python3
"""
Ingest Texas HHSC Assisted Living Facility directory export (Excel) into facilities.

Source: Texas Health and Human Services Commission — Directory of Assisted Living
Facility Providers with an Active License (Excel export).

Upsert conflict target: (state_code, city_slug, slug)

Phase 1 (this script):
  - Optional --metros-only filter (default True): counties in the 7 metro footprints
    defined in the TX Phase 1 plan (harris, dallas, tarrant, bexar, travis/williamson/hays,
    collin, denton).
  - publishable is always false until Phase 2 inspection ingest + recompute.
  - tx_alzheimer_certified = true only when Alzheimer Certificate No is present AND
    Alzheimer Expiration Date is today or later.

Environment:
  DATABASE_URL — required unless --dry-run (dry-run can run without DB for counts only).

Usage:
  python3 scrapers/tx_alf_ingest.py --input ~/Desktop/al.xlsx --dry-run
  python3 scrapers/tx_alf_ingest.py --input ~/Desktop/al.xlsx
  python3 scrapers/tx_alf_ingest.py --input ~/Desktop/al.xlsx --no-metros-only
  python3 scrapers/tx_alf_ingest.py --input ~/Desktop/al.xlsx --metros-only false
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import psycopg
from dotenv import load_dotenv

REPO_ROOT = Path(__file__).resolve().parent.parent
STATE_CODE = "TX"
SCRAPER_NAME = "tx_hhsc_alf_directory"

# Phase 1 — counties included when --metros-only (HHSC prints uppercase names).
METRO_COUNTIES: frozenset[str] = frozenset({
    "HARRIS",
    "FORT BEND",
    "MONTGOMERY",
    "BRAZORIA",
    "GALVESTON",
    "DALLAS",
    "COLLIN",
    "TARRANT",
    "BEXAR",
    "TRAVIS",
    "WILLIAMSON",
    "HAYS",
    "DENTON",
})

# Cities whose slug collides with another state’s hub slug — disambiguate to *-tx.
TX_CITY_SLUG_OVERRIDES: dict[str, str] = {
    "pasadena": "pasadena-tx",
    "arlington": "arlington-tx",
    "paris": "paris-tx",
    "glendale": "glendale-tx",
    "richmond": "richmond-tx",
    "cleveland": "cleveland-tx",
    "springfield": "springfield-tx",
    # HHSC sometimes spells “Sugar Land” as one word
    "sugarland": "sugar-land",
}

HHSC_INFO_BASE = (
    "https://www.hhs.texas.gov/providers/long-term-care-providers/assisted-living"

)


def load_env() -> None:
    for name in (".env.local", ".env"):
        p = REPO_ROOT / name
        if p.is_file():
            load_dotenv(p)


def slugify(text: str) -> str:
    s = text.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "facility"


def titleize(text: str) -> str:
    if not text:
        return text
    stop = {"and", "or", "of", "in", "the", "a", "at", "by", "for", "to",
            "de", "del", "las", "los", "el", "la", "&"}
    words = text.lower().split()
    result = []
    for i, w in enumerate(words):
        result.append(w if (i > 0 and w in stop) else w.capitalize())
    return " ".join(result)


def pad_license_tx(raw: Any) -> str:
    if raw is None or str(raw).strip() == "":
        return "000000"
    try:
        n = int(float(str(raw)))
    except (ValueError, TypeError):
        return "000000"
    return str(n).zfill(6)


def facility_slug(name: str, license_num: str) -> str:
    base = slugify(name)
    suffix = license_num[-6:].lstrip("0") or license_num[-2:]
    return f"{base}-{suffix}"


def slugify_tx_city(raw_city: str) -> str:
    base = slugify(raw_city)
    return TX_CITY_SLUG_OVERRIDES.get(base, base)


def parse_geo(raw: Any) -> tuple[float | None, float | None]:
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s or "," not in s:
        return None, None
    parts = s.split(",")
    try:
        lat = float(parts[0].strip())
        lng = float(parts[1].strip())
        return lat, lng
    except (ValueError, IndexError):
        return None, None


def parse_zip(raw: Any) -> str | None:
    if raw is None or str(raw).strip() == "":
        return None
    s = str(raw).strip()
    if "-" in s:
        s = s.split("-")[0].strip()
    digits = re.sub(r"\D", "", s)
    if len(digits) >= 5:
        return digits[:5]
    if digits:
        return digits.zfill(5)
    return None


def to_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def norm_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def find_header_row(ws: Any) -> tuple[int, list[str]]:
    """Return (0-based row index, header cells)."""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(v == "Facility Name" for v in vals):
            return i, vals
    raise ValueError("Could not find header row containing 'Facility Name'")


def build_header_map(header_cells: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for j, cell in enumerate(header_cells):
        key = norm_header(cell)
        if key:
            idx[key] = j
    return idx


def cell(row: tuple[Any, ...], idx: dict[str, int], *names: str) -> Any:
    for n in names:
        k = norm_header(n)
        if k in idx and idx[k] < len(row):
            return row[idx[k]]
    return None


def load_tx_metro_city_slugs() -> set[str]:
    """Union of citySlugs from the seven TX metro county regions in regions.ts."""
    path = REPO_ROOT / "src" / "lib" / "regions.ts"
    text = path.read_text(encoding="utf-8")
    county_slugs = (
        "harris-county",
        "fort-bend-county",
        "montgomery-county",
        "brazoria-county",
        "galveston-county",
        "dallas-county",
        "collin-county",
        "tarrant-county",
        "denton-county",
        "bexar-county",
        "travis-county",
    )
    out: set[str] = set()
    for cs in county_slugs:
        m = re.search(
            rf'slug:\s*"{re.escape(cs)}"[\s\S]*?citySlugs:\s*\[([\s\S]*?)\]',
            text,
        )
        if not m:
            continue
        out.update(re.findall(r'"([a-z0-9-]+)"', m.group(1)))
    return out


def parse_service_type(raw: Any) -> str | None:
    s = (str(raw or "").strip().upper())
    if "TYPE A" in s:
        return "A"
    if "TYPE B" in s:
        return "B"
    if "TYPE C" in s:
        return "C"
    return None


def build_facility_row(
    row: tuple[Any, ...],
    idx: dict[str, int],
    *,
    phase1_publishable_false: bool,
) -> dict[str, Any] | None:
    raw_name = (cell(row, idx, "Facility Name") or "").strip()
    if not raw_name:
        return None

    lic = pad_license_tx(cell(row, idx, "License No"))
    name = titleize(raw_name) or f"Facility {lic}"
    raw_city = (cell(row, idx, "Physical Address CITY") or "").strip()
    city_slug = slugify_tx_city(raw_city) if raw_city else "unknown-city"

    street = titleize((cell(row, idx, "Physical Address") or "").strip()) or None
    city = titleize(raw_city) if raw_city else None
    zip_str = parse_zip(cell(row, idx, "Physical Address Zipcode"))

    slug = facility_slug(name, lic)

    svc = parse_service_type(cell(row, idx, "Service  Type", "Service Type"))
    tx_license_class = svc

    raw_cap = cell(row, idx, "Alzheimer Capacity")
    alz_cap: int | None = None
    if raw_cap is not None and str(raw_cap).strip() != "":
        try:
            alz_cap = int(float(str(raw_cap)))
        except (ValueError, TypeError):
            alz_cap = None

    cert_no_raw = cell(row, idx, "Alzheimer Certificate No")
    cert_no = (str(cert_no_raw).strip() if cert_no_raw is not None else "") or None
    if cert_no in ("", "None", "0"):
        cert_no = None

    alz_eff = to_date(cell(row, idx, "Alzheimer Certificate Effective Date"))
    alz_exp = to_date(cell(row, idx, "Alzheimer Expiration Date"))

    today = date.today()
    alz_cert_active = bool(
        cert_no
        and alz_exp is not None
        and alz_exp >= today
    )

    lic_eff = to_date(cell(row, idx, "License Effective Date"))
    lic_exp = to_date(cell(row, idx, "License Expiration Date"))
    lic_initial = to_date(cell(row, idx, "Initial License Date"))

    beds_raw = cell(row, idx, "Total Licensed Capacity")
    beds: int | None = None
    if beds_raw is not None and str(beds_raw).strip() != "":
        try:
            beds = int(float(str(beds_raw)))
        except (ValueError, TypeError):
            beds = None

    lat, lng = parse_geo(cell(row, idx, "Geo Location"))

    phone_raw = cell(row, idx, "Facility Phone Number")
    phone = str(phone_raw).strip() if phone_raw else None

    fac_id = cell(row, idx, "Facility ID")
    tx_facility_id = str(fac_id).strip() if fac_id is not None else None

    county = (cell(row, idx, "County") or "").strip().upper() or None
    state_region = (cell(row, idx, "State Region") or "").strip() or None
    suboffice = (cell(row, idx, "HHSC SubOffice") or "").strip() or None

    operator = (cell(row, idx, "Owner_") or "").strip() or None
    if operator:
        operator = titleize(operator)
    mgmt = (cell(row, idx, "Management Company_") or "").strip() or None
    if mgmt:
        mgmt = titleize(mgmt)

    care_category = "alf_memory_care" if alz_cert_active else "alf_general"

    return {
        "state_code": STATE_CODE,
        "name": name,
        "cms_id": None,
        "license_number": lic,
        "license_type": (cell(row, idx, "Service  Type", "Service Type") or "").strip() or None,
        "street": street,
        "city": city,
        "zip": zip_str,
        "city_slug": city_slug,
        "slug": slug,
        "beds": beds,
        "facility_type": "alf",
        "certification_type": "state",
        "operator_name": operator,
        "management_company": mgmt,
        "ownership_type": None,
        "phone": phone,
        "website": None,
        "cms_star_rating": None,
        "last_inspection_date": None,
        "latitude": lat,
        "longitude": lng,
        "source_url": HHSC_INFO_BASE,
        "care_category": care_category,
        "serves_memory_care": alz_cert_active,
        "memory_care_designation": (
            "Texas Alzheimer Certification (HHSC)" if alz_cert_active else None
        ),
        "license_status": "LICENSED",
        "license_expiration": lic_exp,
        "publishable": False if phase1_publishable_false else False,
        "mc_signal_explicit_name": False,
        "mc_signal_chain_name": False,
        "mc_review_status": "auto_published",
        "tx_license_class": tx_license_class,
        "tx_alzheimer_certified": alz_cert_active,
        # HHSC Alzheimer Certification is TX's Tier-1 signal; mirror it into
        # the unified disclosure column so recompute_publishable.py treats
        # certified TX facilities the same as a CA §1569.627 filing.
        "memory_care_disclosure_filed": alz_cert_active,
        "memory_care_disclosure_source": (
            "TX HHSC Alzheimer Certification" if alz_cert_active else None
        ),
        "mc_signal_apfm_listed": False,
        "mc_signal_caring_listed": False,
        "tx_facility_id": tx_facility_id,
        "tx_alzheimer_capacity": alz_cap,
        "tx_alzheimer_cert_no": cert_no,
        "tx_alzheimer_cert_effective": alz_eff,
        "tx_alzheimer_cert_expiration": alz_exp,
        "tx_license_effective": lic_eff,
        "tx_license_expiration": lic_exp,
        "tx_license_initial": lic_initial,
        "tx_state_region": state_region,
        "tx_hhsc_suboffice": suboffice,
        "tx_county": county,
    }


UPSERT_SQL = """
INSERT INTO facilities (
    state_code, name, cms_id,
    license_number, license_type,
    street, city, zip,
    city_slug, slug,
    beds, facility_type, certification_type,
    operator_name, management_company, ownership_type,
    phone, website,
    cms_star_rating, last_inspection_date,
    latitude, longitude,
    source_url,
    care_category, serves_memory_care, memory_care_designation,
    license_status, license_expiration, publishable,
    mc_signal_explicit_name, mc_signal_chain_name, mc_review_status,
    memory_care_disclosure_filed, memory_care_disclosure_source,
    mc_signal_apfm_listed, mc_signal_caring_listed,
    tx_license_class, tx_alzheimer_certified,
    tx_facility_id, tx_alzheimer_capacity, tx_alzheimer_cert_no,
    tx_alzheimer_cert_effective, tx_alzheimer_cert_expiration,
    tx_license_effective, tx_license_expiration, tx_license_initial,
    tx_state_region, tx_hhsc_suboffice, tx_county,
    updated_at
) VALUES (
    %(state_code)s, %(name)s, %(cms_id)s,
    %(license_number)s, %(license_type)s,
    %(street)s, %(city)s, %(zip)s,
    %(city_slug)s, %(slug)s,
    %(beds)s, %(facility_type)s, %(certification_type)s,
    %(operator_name)s, %(management_company)s, %(ownership_type)s,
    %(phone)s, %(website)s,
    %(cms_star_rating)s, %(last_inspection_date)s,
    %(latitude)s, %(longitude)s,
    %(source_url)s,
    %(care_category)s, %(serves_memory_care)s, %(memory_care_designation)s,
    %(license_status)s, %(license_expiration)s, %(publishable)s,
    %(mc_signal_explicit_name)s, %(mc_signal_chain_name)s, %(mc_review_status)s,
    %(memory_care_disclosure_filed)s, %(memory_care_disclosure_source)s,
    %(mc_signal_apfm_listed)s, %(mc_signal_caring_listed)s,
    %(tx_license_class)s, %(tx_alzheimer_certified)s,
    %(tx_facility_id)s, %(tx_alzheimer_capacity)s, %(tx_alzheimer_cert_no)s,
    %(tx_alzheimer_cert_effective)s, %(tx_alzheimer_cert_expiration)s,
    %(tx_license_effective)s, %(tx_license_expiration)s, %(tx_license_initial)s,
    %(tx_state_region)s, %(tx_hhsc_suboffice)s, %(tx_county)s,
    now()
)
ON CONFLICT (state_code, city_slug, slug) DO UPDATE SET
    name                         = EXCLUDED.name,
    license_number               = EXCLUDED.license_number,
    license_type                 = EXCLUDED.license_type,
    street                       = EXCLUDED.street,
    city                         = EXCLUDED.city,
    zip                          = EXCLUDED.zip,
    beds                         = EXCLUDED.beds,
    facility_type                = EXCLUDED.facility_type,
    certification_type           = EXCLUDED.certification_type,
    operator_name                = EXCLUDED.operator_name,
    management_company           = EXCLUDED.management_company,
    phone                        = EXCLUDED.phone,
    last_inspection_date         = EXCLUDED.last_inspection_date,
    latitude                     = EXCLUDED.latitude,
    longitude                    = EXCLUDED.longitude,
    source_url                   = EXCLUDED.source_url,
    care_category                = EXCLUDED.care_category,
    serves_memory_care           = EXCLUDED.serves_memory_care,
    memory_care_designation      = EXCLUDED.memory_care_designation,
    license_status               = EXCLUDED.license_status,
    license_expiration           = EXCLUDED.license_expiration,
    publishable                  = EXCLUDED.publishable,
    mc_signal_explicit_name      = EXCLUDED.mc_signal_explicit_name,
    mc_signal_chain_name         = EXCLUDED.mc_signal_chain_name,
    mc_review_status             = CASE
        WHEN facilities.mc_review_status IN ('reviewed_publish','reviewed_reject')
        THEN facilities.mc_review_status
        ELSE EXCLUDED.mc_review_status
    END,
    tx_license_class             = EXCLUDED.tx_license_class,
    tx_alzheimer_certified       = EXCLUDED.tx_alzheimer_certified,
    tx_facility_id               = EXCLUDED.tx_facility_id,
    tx_alzheimer_capacity        = EXCLUDED.tx_alzheimer_capacity,
    tx_alzheimer_cert_no         = EXCLUDED.tx_alzheimer_cert_no,
    tx_alzheimer_cert_effective  = EXCLUDED.tx_alzheimer_cert_effective,
    tx_alzheimer_cert_expiration = EXCLUDED.tx_alzheimer_cert_expiration,
    tx_license_effective         = EXCLUDED.tx_license_effective,
    tx_license_expiration        = EXCLUDED.tx_license_expiration,
    tx_license_initial           = EXCLUDED.tx_license_initial,
    tx_state_region              = EXCLUDED.tx_state_region,
    tx_hhsc_suboffice            = EXCLUDED.tx_hhsc_suboffice,
    tx_county                    = EXCLUDED.tx_county,
    updated_at                   = now()
"""


def upsert_facilities(conn: psycopg.Connection, rows: list[dict[str, Any]]) -> None:
    """Batch upsert — psycopg3 executemany pipelines statements to the server."""
    if not rows:
        return
    with conn.cursor() as cur:
        cur.executemany(UPSERT_SQL, rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest HHSC ALF directory Excel → facilities")
    parser.add_argument("--input", type=Path, required=True, help="Path to HHSC ALF Excel export")
    parser.add_argument("--dry-run", action="store_true", help="Print stats; optional DB skip")
    parser.add_argument(
        "--metros-only",
        dest="metros_only",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Only counties in Phase 1 metro footprint (default: true)",
    )
    parser.add_argument("--limit", type=int, default=None, help="Process at most N data rows")
    args = parser.parse_args()

    load_env()

    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 2

    if not args.input.is_file():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        return 1

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row_idx, header_cells = find_header_row(ws)
    header_map = build_header_map(header_cells)

    allowed_slugs = load_tx_metro_city_slugs()
    rows_iter = ws.iter_rows(
        min_row=header_row_idx + 2,
        values_only=True,
    )

    total_read = 0
    filtered_out = 0
    built: list[dict[str, Any]] = []
    orphan_city: Counter[str] = Counter()
    alz_active_count = 0

    for raw_row in rows_iter:
        if args.limit is not None and len(built) >= args.limit:
            break
        if not raw_row or not any(raw_row):
            continue
        total_read += 1

        county = (cell(raw_row, header_map, "County") or "").strip().upper()
        if args.metros_only and county not in METRO_COUNTIES:
            filtered_out += 1
            continue

        rec = build_facility_row(
            raw_row,
            header_map,
            phase1_publishable_false=True,
        )
        if rec is None:
            continue

        if rec["tx_alzheimer_certified"]:
            alz_active_count += 1

        cs = rec["city_slug"]
        if cs not in allowed_slugs:
            orphan_city[rec["city"] or cs] += 1

        built.append(rec)

    wb.close()

    print(f"Rows read (data): {total_read}")
    print(f"Filtered out (non-metro county): {filtered_out}")
    print(f"Rows after filter + parse: {len(built)}")
    print(f"Alzheimer-certified (non-expired cert): {alz_active_count}")

    if orphan_city:
        print("\nCity names whose slug is not in TX metro citySlugs (top 25):")
        for city, n in orphan_city.most_common(25):
            print(f"  {city!r}: {n}")

    if args.dry_run:
        print("\nDry-run complete — no database writes.")
        return 0

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("DATABASE_URL not set; cannot upsert.", file=sys.stderr)
        return 1

    with psycopg.connect(dsn) as conn:
        upsert_facilities(conn, built)
        conn.commit()

    print(f"\nUpsert complete: {len(built)} rows (batched upsert)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
