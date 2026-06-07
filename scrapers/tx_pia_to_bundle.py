#!/usr/bin/env python3
"""
Convert HHSC PIA (Public Information Act) extracts → format_version: 1 bundle
for [`scrapers/tx_inspections_ingest.py`](./tx_inspections_ingest.py)
`--import-json`.

Why this exists
---------------
TULIP's interactive Aura capture path (see [`tx_tulip_to_bundle.py`](./tx_tulip_to_bundle.py))
is fine for tens of facilities, but bulk coverage of all ~515 Alzheimer-certified
TX ALFs requires the agency's structured records. HHSC fulfils PIA requests
in two distinct shapes — neither contains inspector-narrative text by default;
narratives must be requested per-event in a follow-up records request.

Supported input shapes
----------------------
**Bulk PIA (2026-05 fulfilment) — two sheets, no narratives:**

1. **FVH** (Facility Visit History): `CARES Region, ABBREV, STATEID, NAME, Address,
   City, Zip, EVENTID, EntranceDate, EXIT_DATE, VisitType, VisitSequence, FullBook,
   FullBookType, Investigation, InfectionControl, SurveyLocation, StateViolations`.
   One row per visit; `StateViolations` is a coded string like
   `P-0031 s/s: 0; P-2554 s/s: F` — rule code + CMS A-L scope/severity letter.
   Narrative: not included; request per-EVENTID via follow-up.

2. **IntakeHistory** (complaints + incidents): `Facility ID, Agency/Facility, Program
   Type, Agency/Facility Address, RS Case No., Case Type, Priority, Date Received,
   Date/Time Sent to Region, Status, Allegation, Findings`. One row per complaint;
   `Findings` is `UNSUBSTANTIATED`, `SUBSTANTIATED AND CITED`, `SUBSTANTIATED BUT NOT
   CITED`, etc. Narrative: not included; request per-RS-Case-No via follow-up.

3. **Legacy / TULIP-style** (single sheet with narratives): `License No, Exit Date,
   State Violation Cited, ...` — the original PIA shape this parser was built for.
   Still supported via the same CLI; column matching is fuzzy.

Roster lookup (FVH + IntakeHistory only)
----------------------------------------
FVH and IntakeHistory join by `STATEID` / `Facility ID` (HHSC's internal facility ID),
not by license number. The parser needs to resolve those IDs to license numbers so
the existing ingest pipeline (which keys on license_number) works unchanged.

Generate the lookup once with [`tx_export_facility_lookup.py`](./tx_export_facility_lookup.py):

    python3 scrapers/tx_export_facility_lookup.py
    # writes scrapers/data/tx_facility_lookup.csv (gitignored)

If the lookup is missing, FVH/Intake sheets are skipped with a warning.
Legacy single-sheet PIA inputs do not need the lookup.

Usage
-----
    # Bulk PIA (the 2026-05 HHSC delivery shape)
    python3 scrapers/tx_pia_to_bundle.py \\
      --input .firecrawl/tx-pia/2026-05-bulk/FVH_AllALFs_05062020_05062026.xlsx \\
      --input .firecrawl/tx-pia/2026-05-bulk/IntakeHistory_AllALFs_05062020_05062026.xlsx \\
      --output .firecrawl/tx-pia/2026-05-bulk/bundle.json

    # Legacy single-sheet PIA / TULIP-style with narratives
    python3 scrapers/tx_pia_to_bundle.py \\
      --input .firecrawl/tx-pia/2026-05-pia-response.xlsx \\
      --output .firecrawl/tx-pia/bundle.json

    # Dry-run (parse + summarize, do not write bundle)
    python3 scrapers/tx_pia_to_bundle.py --input pia.xlsx --dry-run

CMS scope/severity grid (A-L)
-----------------------------
Letters in `StateViolations` are CMS-standard. Persisted verbatim in
`deficiencies.state_severity_raw`. `immediate_jeopardy = True` is set for
J / K / L per CMS definition. The numeric `severity` (1-4) is left null;
mapping the 12-cell grid onto the 4-rank ordinal is a methodology call,
not an ingest concern.

Source URL
----------
PIA-sourced inspections do not have a public detail URL. We synthesize:
    https://hhs.texas.gov/ltcr-pia/<license>#<inspection_date>           (visits)
    https://hhs.texas.gov/ltcr-pia/<license>#complaint-<rs_case_no>      (intake)
captured as `inspections.source_url`. The actual fulfilment file should be
archived separately under `.firecrawl/tx-pia/`.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

# ─── Constants ────────────────────────────────────────────────────────────────

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_ROSTER = Path(__file__).resolve().parent / "data" / "tx_facility_lookup.csv"

VISIT_URL_TEMPLATE = "https://hhs.texas.gov/ltcr-pia/{license}#{date}"
COMPLAINT_URL_TEMPLATE = "https://hhs.texas.gov/ltcr-pia/{license}#complaint-{rs}"

# CMS scope/severity letters that count as immediate jeopardy.
IJ_LETTERS = {"J", "K", "L"}

# ─── Column detection (legacy / fuzzy path) ──────────────────────────────────

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "license_number": (
        "license", "licenseno", "licensenumber", "licensenum", "lic",
        "facilitylicense", "alflicense",
        "hfid", "healthfacilityid",
    ),
    "facility_id_tx": (
        "stateid", "facilityid", "providerid", "facilityidnumber",
    ),
    "facility_name": (
        "facilityname", "agencyfacility", "provider", "providername", "name",
    ),
    "inspection_date": (
        "exitdate", "visitexitdate", "surveyexit", "surveyexitdate",
        "inspectiondate", "visitdate", "eventdate",
    ),
    "incident_date": (
        "incidentdate", "complaintreceived", "complaintreceiveddate",
        "complaintdate", "intakedate", "datereceived",
    ),
    "visit_type": (
        "visittype", "visitsurveytype", "surveytype", "eventtype",
        "visitcategory", "category", "casetype",
    ),
    "citation": (
        "violationcited", "stateviolationcited", "citationnarrative",
        "narrative", "violation", "finding", "deficiencytext",
    ),
    "citation_code": (
        "citationnumber", "rulecitation", "reference", "citation",
        "rule", "code", "referencerule",
    ),
    "corrected_date": (
        "correcteddate", "planofcorrectiondate", "pocdate",
        "datecorrected", "correctionsdate",
    ),
    "complaint_id": (
        "complaintnumber", "complaintid", "intakenumber", "intake",
        "rscaseno", "rscaseno.", "rscasenumber", "rscaseid",
    ),
    "event_id": (
        "surveyid", "eventid", "visitid", "inspectionid", "trackingid",
    ),
    "outcome": (
        "findings", "disposition", "result",
    ),
    "allegation": (
        "allegation", "allegationcategory",
    ),
}


def _normalize_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def detect_columns(headers: list[str]) -> dict[str, str]:
    """
    Map our canonical field names → actual sheet column names. Picks the
    first header whose normalized form contains the alias keyword (in order
    of specificity).
    """
    normalized = {h: _normalize_header(h) for h in headers}
    out: dict[str, str] = {}
    for canon, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for raw, norm in normalized.items():
                if alias in norm:
                    out.setdefault(canon, raw)
                    break
            if canon in out:
                break
    return out


# ─── Date / value helpers ────────────────────────────────────────────────────

DATE_FORMATS = (
    "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y",
    "%Y/%m/%d", "%d-%b-%Y", "%d/%m/%Y",
)


def parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()[:10]
        except Exception:  # noqa: BLE001
            pass
    s = str(value).strip()
    if not s:
        return None
    # Trim a trailing time suffix (e.g. "12/5/2025 3:56 PM").
    s_date_only = s.split(" ")[0]
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s_date_only, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def s(value: Any) -> str | None:
    if value is None:
        return None
    txt = str(value).strip()
    return txt or None


def pad_license(value: Any) -> str | None:
    raw = s(value)
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return raw
    # TX licenses are 6-digit; MN HFID (and similar) are longer — pad to 10 when needed.
    if len(digits) > 6:
        return digits.zfill(10)
    return digits.zfill(6)


def pad_facility_id_tx(value: Any) -> str | None:
    """Mirror the roster format: 6-digit zero-padded HHSC facility ID."""
    raw = s(value)
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return None
    return digits.zfill(6)


# ─── State Violations parser (FVH-only) ──────────────────────────────────────

# Matches `P-1972 s/s: E` or `P-12345 s/s: 0` etc. Letters A-L plus 0 (no finding).
SV_PATTERN = re.compile(r"P-(\d+)\s*s/s\s*:\s*([A-L0])", re.IGNORECASE)


def parse_state_violations(text: Any) -> list[tuple[str, str]]:
    """
    Parse the FVH StateViolations cell into [(code, ss_letter), ...].
    Skips chunks with `s/s: 0` (no finding for that rule). Returns [] on
    empty input or `(none)` markers.
    """
    if text is None:
        return []
    raw = str(text).strip()
    if not raw or raw.lower() in ("(none)", "none"):
        return []
    out: list[tuple[str, str]] = []
    for m in SV_PATTERN.finditer(raw):
        code = f"P-{m.group(1)}"
        letter = m.group(2).upper()
        if letter == "0":
            continue
        out.append((code, letter))
    return out


# ─── Roster loading ──────────────────────────────────────────────────────────

def load_tx_roster(path: Path | None) -> dict[str, dict[str, Any]]:
    """
    Load the TX facility lookup CSV (see tx_export_facility_lookup.py).
    Returns {tx_facility_id_padded: {license_number, name, city, county,
    tx_alzheimer_certified}}.
    """
    if path is None or not path.is_file():
        return {}
    out: dict[str, dict[str, Any]] = {}
    with path.open(encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            fid_padded = pad_facility_id_tx(row.get("tx_facility_id"))
            if not fid_padded:
                continue
            out[fid_padded] = {
                "license_number": pad_license(row.get("license_number")),
                "tx_facility_id": fid_padded,
                "name": row.get("name"),
                "city": row.get("city"),
                "tx_county": row.get("tx_county"),
                "tx_alzheimer_certified": row.get("tx_alzheimer_certified") == "True",
            }
    return out


# ─── Sheet readers ───────────────────────────────────────────────────────────

def read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]
    return headers, rows


def read_xlsx(path: Path) -> list[tuple[str, list[str], list[dict[str, Any]]]]:
    """Returns [(sheet_name, headers, rows)] across all sheets."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        print(
            "ERROR: openpyxl not installed. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(2)

    wb = load_workbook(path, data_only=True, read_only=True)
    out = []
    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        rows = []
        for r in rows_iter:
            if all(v is None or v == "" for v in r):
                continue
            rows.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
        if rows:
            out.append((sheet.title, headers, rows))
    return out


def read_input(path: Path) -> list[tuple[str, list[str], list[dict[str, Any]]]]:
    """Returns [(sheet_label, headers, rows)] for any supported file type."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    if suffix == ".csv":
        h, r = read_csv(path)
        return [(path.name, h, r)]
    if suffix in (".tsv", ".txt"):
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            return [(path.name, list(reader.fieldnames or []), [dict(r) for r in reader])]
    raise ValueError(f"Unsupported file type: {path.suffix} ({path})")


# ─── Sheet shape detection ───────────────────────────────────────────────────

def detect_sheet_shape(headers: list[str]) -> str:
    """Return one of: 'fvh', 'intake', 'legacy'."""
    norm = {_normalize_header(h) for h in headers}
    has_eventid = any("eventid" in h for h in norm)
    has_state_violations = any("stateviolation" in h for h in norm)
    has_exit_date = any("exitdate" in h or "exit_date" in h for h in norm)
    if has_eventid and has_state_violations and has_exit_date:
        return "fvh"
    has_rs_case = any("rscaseno" in h for h in norm)
    has_allegation = any("allegation" in h for h in norm)
    has_findings = any(h == "findings" for h in norm)
    if has_rs_case and has_allegation and has_findings:
        return "intake"
    return "legacy"


# ─── Per-shape processors ────────────────────────────────────────────────────

def _process_fvh_sheet(
    rows: list[dict[str, Any]],
    headers: list[str],
    roster: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    FVH: one inspection per row (keyed by EVENTID). Deficiencies are parsed from
    the StateViolations coded string. Returns (inspections_grouped_by_license,
    stats).

    Shape of each emitted inspection block:
        {
          "license_number": "<6-digit padded>",
          "inspections": [
            {
              "inspection_date": "YYYY-MM-DD",
              "inspection_type": "complaint" | "standard",
              "is_complaint": bool,
              "source_url": "...",
              "raw_data": {pia_source, fvh_event_id, visit_type, ...},
              "deficiencies": [{code, state_severity_raw, immediate_jeopardy, ...}, ...],
            }, ...
          ]
        }
    """
    # Map FVH-fixed columns by name (these are stable in HHSC's bulk delivery).
    H = {h: h for h in headers}
    def col(name: str) -> str | None:
        return H.get(name)

    by_license: dict[str, list[dict[str, Any]]] = defaultdict(list)
    stats = {
        "rows": 0, "skipped_no_stateid": 0, "skipped_unmatched_facility": 0,
        "skipped_no_date": 0, "inspections": 0, "deficiencies": 0,
        "inspections_with_violations": 0,
    }

    for row in rows:
        stats["rows"] += 1
        stateid_padded = pad_facility_id_tx(row.get("STATEID"))
        if not stateid_padded:
            stats["skipped_no_stateid"] += 1
            continue
        fac = roster.get(stateid_padded)
        if not fac:
            stats["skipped_unmatched_facility"] += 1
            continue
        license_no = fac["license_number"]
        if not license_no:
            stats["skipped_unmatched_facility"] += 1
            continue
        exit_date = parse_date(row.get("EXIT_DATE"))
        if not exit_date:
            stats["skipped_no_date"] += 1
            continue

        violations = parse_state_violations(row.get("StateViolations"))
        deficiencies: list[dict[str, Any]] = []
        for code, letter in violations:
            deficiencies.append({
                "code": code,
                "state_severity_raw": letter,
                "scope_severity_code": letter,
                "description": None,
                "inspector_narrative": None,
                "cited_date": exit_date,
                "corrected_date": None,
                "immediate_jeopardy": letter in IJ_LETTERS,
            })

        visit_type = s(row.get("VisitType")) or "Unknown"
        investigation = s(row.get("Investigation"))
        is_complaint = bool(investigation and investigation.lower() == "yes")

        inspection_block = {
            "inspection_date": exit_date,
            "inspection_type": "complaint" if is_complaint else "standard",
            "is_complaint": is_complaint,
            "complaint_id": None,
            "source_url": VISIT_URL_TEMPLATE.format(license=license_no, date=exit_date),
            "deficiencies": deficiencies,
            "raw_data": {
                "pia_source": True,
                "fvh_event_id": s(row.get("EVENTID")),
                "tx_facility_id": stateid_padded,
                "visit_type": visit_type,
                "visit_sequence": s(row.get("VisitSequence")),
                "full_book": s(row.get("FullBook")),
                "full_book_type": s(row.get("FullBookType")),
                "investigation": investigation,
                "infection_control": s(row.get("InfectionControl")),
                "survey_location": s(row.get("SurveyLocation")),
                "cares_region": s(row.get("CARES Region")),
                "raw_state_violations": s(row.get("StateViolations")),
                "facility_name_at_visit": s(row.get("NAME")),
                "address_at_visit": s(row.get("Address")),
            },
        }
        by_license[license_no].append(inspection_block)
        stats["inspections"] += 1
        stats["deficiencies"] += len(deficiencies)
        if deficiencies:
            stats["inspections_with_violations"] += 1

    return [
        {"license_number": lic, "inspections": insps}
        for lic, insps in sorted(by_license.items())
    ], stats


def _process_intake_sheet(
    rows: list[dict[str, Any]],
    headers: list[str],
    roster: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    IntakeHistory: one complaint inspection per row, keyed by RS Case No.
    Deficiency synthesized only when Findings starts with "SUBSTANTIATED AND CITED";
    "SUBSTANTIATED BUT NOT CITED" is recorded on the inspection but produces no
    deficiency row (no formal citation issued).
    """
    by_license: dict[str, list[dict[str, Any]]] = defaultdict(list)
    stats = {
        "rows": 0, "skipped_no_facid": 0, "skipped_unmatched_facility": 0,
        "skipped_no_date": 0, "inspections": 0, "deficiencies": 0,
        "substantiated_and_cited": 0, "substantiated_not_cited": 0,
        "unsubstantiated": 0, "other_outcome": 0,
    }

    for row in rows:
        stats["rows"] += 1
        fac_id_padded = pad_facility_id_tx(row.get("Facility ID"))
        if not fac_id_padded:
            stats["skipped_no_facid"] += 1
            continue
        fac = roster.get(fac_id_padded)
        if not fac:
            stats["skipped_unmatched_facility"] += 1
            continue
        license_no = fac["license_number"]
        if not license_no:
            stats["skipped_unmatched_facility"] += 1
            continue
        date_received = parse_date(row.get("Date Received"))
        if not date_received:
            stats["skipped_no_date"] += 1
            continue

        rs_case_no = s(row.get("RS Case No.")) or "unknown"
        findings = s(row.get("Findings")) or ""
        findings_upper = findings.upper()
        allegation = s(row.get("Allegation"))

        deficiencies: list[dict[str, Any]] = []
        if findings_upper.startswith("SUBSTANTIATED AND CITED"):
            stats["substantiated_and_cited"] += 1
            deficiencies.append({
                "code": f"intake-{rs_case_no}",
                "category": allegation,
                "state_severity_raw": findings,
                "description": None,
                "inspector_narrative": None,
                "cited_date": date_received,
                "corrected_date": None,
                "immediate_jeopardy": False,
            })
        elif findings_upper.startswith("SUBSTANTIATED"):
            stats["substantiated_not_cited"] += 1
        elif findings_upper.startswith("UNSUBSTANTIATED"):
            stats["unsubstantiated"] += 1
        else:
            stats["other_outcome"] += 1

        inspection_block = {
            "inspection_date": date_received,
            "incident_date": date_received,
            "inspection_type": "complaint",
            "is_complaint": True,
            "complaint_id": rs_case_no,
            "source_url": COMPLAINT_URL_TEMPLATE.format(license=license_no, rs=rs_case_no),
            "outcome": findings,
            "deficiencies": deficiencies,
            "raw_data": {
                "pia_source": True,
                "intake_source": True,
                "tx_facility_id": fac_id_padded,
                "rs_case_no": rs_case_no,
                "case_type": s(row.get("Case Type")),
                "priority": s(row.get("Priority")),
                "status": s(row.get("Status")),
                "allegation": allegation,
                "findings": findings,
                "date_sent_to_region": s(row.get("Date/Time Sent to Region")),
                "facility_name_at_intake": s(row.get("Agency/Facility")),
                "address_at_intake": s(row.get("Agency/Facility Address")),
            },
        }
        by_license[license_no].append(inspection_block)
        stats["inspections"] += 1
        stats["deficiencies"] += len(deficiencies)

    return [
        {"license_number": lic, "inspections": insps}
        for lic, insps in sorted(by_license.items())
    ], stats


def _process_legacy_sheet(
    rows: list[dict[str, Any]],
    headers: list[str],
) -> tuple[dict[tuple[str, str], list[tuple[dict[str, Any], dict[str, Any]]]], dict[str, int]]:
    """
    Legacy / TULIP-style PIA: fuzzy column matching, group rows by
    (license, inspection_date), one inspection per group. Returns the
    grouped intermediate (caller assembles the final bundle).
    """
    cols = detect_columns(headers)
    stats = {"rows": 0, "skipped": 0}
    groups: dict[tuple[str, str], list[tuple[dict[str, Any], dict[str, Any]]]] = defaultdict(list)
    if "license_number" not in cols or "inspection_date" not in cols:
        stats["rows"] = len(rows)
        stats["skipped"] = len(rows)
        return groups, stats
    for row in rows:
        stats["rows"] += 1
        canon = {canon_key: row.get(src) for canon_key, src in cols.items()}
        lic = pad_license(canon.get("license_number"))
        insp = parse_date(canon.get("inspection_date"))
        if not lic or not insp:
            stats["skipped"] += 1
            continue
        groups[(lic, insp)].append((canon, row))
    return groups, stats


def _legacy_groups_to_facilities(
    groups: dict[tuple[str, str], list[tuple[dict[str, Any], dict[str, Any]]]]
) -> list[dict[str, Any]]:
    """Convert legacy grouped rows → facility blocks compatible with the bundle."""
    facilities: dict[str, dict[str, Any]] = defaultdict(lambda: {"inspections": []})
    for (lic, insp_date), entries in groups.items():
        incident_date = next(
            (parse_date(c.get("incident_date")) for c, _ in entries if c.get("incident_date")),
            None,
        )
        complaint_id = next(
            (s(c.get("complaint_id")) for c, _ in entries if c.get("complaint_id")),
            None,
        )
        visit_types = [s(c.get("visit_type")) for c, _ in entries if c.get("visit_type")]
        primary_visit_type = visit_types[0] if visit_types else None
        is_complaint = bool(complaint_id) or any(
            "complaint" in (vt or "").lower() for vt in visit_types
        )
        deficiencies = []
        for canon, _raw in entries:
            cit_text = s(canon.get("citation"))
            cit_code = s(canon.get("citation_code"))
            if not cit_text and not cit_code:
                continue
            deficiencies.append({
                "code": cit_code,
                "state_severity_raw": s(canon.get("visit_type")) or primary_visit_type,
                "description": cit_text,
                "inspector_narrative": cit_text,
                "cited_date": insp_date,
                "corrected_date": parse_date(canon.get("corrected_date")),
                "immediate_jeopardy": False,
            })
        facilities[lic]["license_number"] = lic
        facilities[lic]["inspections"].append({
            "inspection_date": insp_date,
            "incident_date": incident_date,
            "inspection_type": "complaint" if is_complaint else "comprehensive",
            "is_complaint": is_complaint,
            "complaint_id": complaint_id,
            "source_url": VISIT_URL_TEMPLATE.format(license=lic, date=insp_date),
            "raw_data": {
                "pia_source": True,
                "visit_type": primary_visit_type,
                "rows": [r for _, r in entries],
            },
            "deficiencies": deficiencies,
        })
    return [
        {"license_number": lic, "inspections": data["inspections"]}
        for lic, data in sorted(facilities.items())
    ]


# ─── Bundle assembly ─────────────────────────────────────────────────────────

def build_bundle(
    sheets: list[tuple[str, list[str], list[dict[str, Any]]]],
    roster: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """
    Process each sheet according to its detected shape and merge into a single
    format_version: 1 bundle. Returns (bundle, summary_stats).
    """
    by_license_aggregated: dict[str, list[dict[str, Any]]] = defaultdict(list)
    summary: dict[str, Any] = {"sheets": []}

    for sheet_name, headers, rows in sheets:
        shape = detect_sheet_shape(headers)
        sheet_stat: dict[str, Any] = {"name": sheet_name, "shape": shape, "rows": len(rows)}

        if shape == "fvh":
            if not roster:
                sheet_stat["error"] = "no roster lookup; FVH sheet skipped"
                summary["sheets"].append(sheet_stat)
                continue
            facility_blocks, st = _process_fvh_sheet(rows, headers, roster)
            sheet_stat.update(st)
            for blk in facility_blocks:
                by_license_aggregated[blk["license_number"]].extend(blk["inspections"])

        elif shape == "intake":
            if not roster:
                sheet_stat["error"] = "no roster lookup; IntakeHistory sheet skipped"
                summary["sheets"].append(sheet_stat)
                continue
            facility_blocks, st = _process_intake_sheet(rows, headers, roster)
            sheet_stat.update(st)
            for blk in facility_blocks:
                by_license_aggregated[blk["license_number"]].extend(blk["inspections"])

        else:
            groups, st = _process_legacy_sheet(rows, headers)
            sheet_stat.update(st)
            facility_blocks = _legacy_groups_to_facilities(groups)
            for blk in facility_blocks:
                by_license_aggregated[blk["license_number"]].extend(blk["inspections"])

        summary["sheets"].append(sheet_stat)

    bundle = {
        "format_version": 1,
        "facilities": [
            {"license_number": lic, "inspections": insps}
            for lic, insps in sorted(by_license_aggregated.items())
        ],
    }
    summary["facilities"] = len(bundle["facilities"])
    summary["inspections"] = sum(len(f["inspections"]) for f in bundle["facilities"])
    summary["deficiencies"] = sum(
        len(i.get("deficiencies", []))
        for f in bundle["facilities"]
        for i in f["inspections"]
    )
    return bundle, summary


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="HHSC PIA extract → ingest bundle")
    ap.add_argument(
        "--input", action="append", required=True,
        help="Input PIA file (CSV / TSV / XLSX). Repeat to merge multiple files.",
    )
    ap.add_argument(
        "--output", default=None,
        help="Bundle JSON output path. Required unless --dry-run.",
    )
    ap.add_argument(
        "--roster", type=Path, default=DEFAULT_ROSTER,
        help=f"TX facility lookup CSV (default: {DEFAULT_ROSTER}).",
    )
    ap.add_argument(
        "--dry-run", action="store_true",
        help="Parse + summarize without writing the bundle.",
    )
    args = ap.parse_args()

    sheets: list[tuple[str, list[str], list[dict[str, Any]]]] = []
    for raw in args.input:
        path = Path(raw)
        if not path.exists():
            print(f"ERROR: input not found: {path}", file=sys.stderr)
            sys.exit(1)
        sheets.extend(read_input(path))

    if not sheets:
        print("ERROR: no readable sheets in input(s)", file=sys.stderr)
        sys.exit(1)

    roster = load_tx_roster(args.roster)
    if not roster:
        print(
            f"  [warn] no roster found at {args.roster}. FVH/Intake sheets will "
            f"be skipped. Generate with `python3 scrapers/tx_export_facility_lookup.py`.",
            file=sys.stderr,
        )

    print(f"Loaded {len(roster)} TX facilities into roster lookup.")
    print("Sheets:")
    for name, headers, rows in sheets:
        shape = detect_sheet_shape(headers)
        print(f"  - {name}: shape={shape}, {len(rows)} rows")

    bundle, summary = build_bundle(sheets, roster)

    print("\nBundle summary:")
    for sheet_stat in summary["sheets"]:
        print(f"  - sheet '{sheet_stat['name']}' ({sheet_stat['shape']}):")
        for k, v in sheet_stat.items():
            if k in ("name", "shape"):
                continue
            print(f"      {k}: {v}")
    print(f"\n  Total facilities: {summary['facilities']}")
    print(f"  Total inspections: {summary['inspections']}")
    print(f"  Total deficiencies: {summary['deficiencies']}")

    if args.dry_run:
        print("(dry-run: bundle not written)")
        return

    if not args.output:
        print("ERROR: --output is required unless --dry-run", file=sys.stderr)
        sys.exit(1)

    Path(args.output).write_text(json.dumps(bundle, indent=2), encoding="utf-8")
    print(f"\nWrote {args.output}")


if __name__ == "__main__":
    main()
